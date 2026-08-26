"""Create deterministic repeated k-fold lists for IPV training and validation."""

import csv
import random
import re
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .fold_utils import natural_key


NUM_REPETITIONS = 3
NUM_FOLDS_PER_REPETITION = 5
BASE_SEED = 42
TEST_SAMPLE_IDS = ['A303', 'A275', 'A270', 'A268', 'A259', 'A258', 'A257', 'A246', 'A243', 'A237',
                   'A235', 'A296', 'A230', 'A225', 'A222', 'A221', 'A217', 'A215', 'A242', 'A207']

MARK_LIST_PATH = Path(r'C:\Storage\Datasets\IPV\OriginalData\doctors_resampled_transverseMarkList.txt')
OUTPUT_DIR = Path(r'C:\Storage\Datasets\IPV\OriginalData\folds_network_study')

CLEAN_OUTPUT_DIR = False
SORT_OUTPUT_FILES = True
WRITE_SUMMARY_CSV = True
WRITE_MEMBERSHIP_CSV = True

SUMMARY_FILE_NAME = 'repeated_kfold_summary.csv'
MEMBERSHIP_FILE_NAME = 'repeated_kfold_membership.csv'
TEST_CASES_WORKBOOK_NAME = 'test_cases.xlsx'
TEST_CASES_SHEET_NAME = 'test_cases'
REPETITION_DIR_PATTERN = re.compile(r'^repetition_(\d+)$')
LEGACY_FOLD_FILE_PATTERN = re.compile(r'^[A-Za-z]+_f\d+\.txt$')
LEGACY_SUMMARY_FILE_NAMES = ('fold_summary.csv', 'fold_membership.csv')


def read_sample_ids(mark_list_path):
    """Read sample stems from the first column of a mark-list file."""
    mark_list_path = Path(mark_list_path)

    if not mark_list_path.is_file():
        raise FileNotFoundError(f'Mark-list file not found: {mark_list_path}')

    return [Path(line.split()[0]).stem for line in mark_list_path.read_text(encoding='utf-8').splitlines() if line.strip()]


def check_duplicates(sample_ids, source_name):
    """Reject duplicate sample IDs."""
    seen = set()
    duplicates = set()

    for sample_id in sample_ids:
        if sample_id in seen:
            duplicates.add(sample_id)
        seen.add(sample_id)

    if duplicates:
        raise ValueError(f'Duplicate sample IDs found in {source_name}: {", ".join(sorted(duplicates, key=natural_key))}')


def load_unique_sample_ids(mark_list_path):
    """Load, validate, and naturally sort all sample IDs."""
    sample_ids = read_sample_ids(mark_list_path)
    check_duplicates(sample_ids, str(mark_list_path))

    if not sample_ids:
        raise ValueError(f'No sample IDs found in {mark_list_path}')

    return sorted(sample_ids, key=natural_key)


def normalise_test_sample_ids(test_sample_ids, all_sample_ids, mark_list_path):
    """Validate optional fixed held-out test IDs."""
    if test_sample_ids is None:
        return []

    if isinstance(test_sample_ids, (str, Path)):
        raise ValueError('TEST_SAMPLE_IDS must be an iterable of sample IDs, not one bare string or path.')

    try:
        raw_ids = list(test_sample_ids)
    except TypeError as error:
        raise ValueError('TEST_SAMPLE_IDS must be None or an iterable of sample ID strings.') from error

    available = set(all_sample_ids)
    normalised = []

    for index, raw_id in enumerate(raw_ids, start=1):
        if not isinstance(raw_id, (str, Path)):
            raise ValueError(f'TEST_SAMPLE_IDS entry {index} must be a string or path. Got: {raw_id!r}')

        text = str(raw_id).strip()

        if not text:
            raise ValueError(f'TEST_SAMPLE_IDS entry {index} is blank.')

        normalised.append(text if text in available else Path(text).stem)

    check_duplicates(normalised, 'TEST_SAMPLE_IDS')
    unknown = sorted(set(normalised) - available, key=natural_key)

    if unknown:
        raise ValueError(f'Test sample IDs were not found in mark-list file {mark_list_path}: {", ".join(unknown)}')

    return sorted(normalised, key=natural_key)


def validate_generation_args(sample_count, num_repetitions, num_folds):
    """Validate repeated k-fold counts."""
    if int(num_repetitions) < 1:
        raise ValueError(f'num_repetitions must be at least 1. Got: {num_repetitions}')
    if int(num_folds) < 2:
        raise ValueError(f'num_folds must be at least 2 for k-fold cross-validation. Got: {num_folds}')
    if int(sample_count) < int(num_folds):
        raise ValueError(f'Not enough samples for {num_folds} folds. Found {sample_count}; at least {num_folds} are required.')


def create_kfold_splits(shuffled_ids, num_folds):
    """Create balanced validation folds and complementary training sets."""
    validation_chunks = [[] for _ in range(int(num_folds))]

    for index, sample_id in enumerate(shuffled_ids):
        validation_chunks[index % int(num_folds)].append(sample_id)

    folds = []

    for fold, validation_ids in enumerate(validation_chunks, start=1):
        validation_set = set(validation_ids)
        folds.append({'fold': fold,
                      'training': [sample_id for sample_id in shuffled_ids if sample_id not in validation_set],
                      'validation': validation_ids})

    return folds


def validate_kfold_splits(folds, all_sample_ids, repetition):
    """Require a complete, non-overlapping k-fold repetition."""
    all_samples = set(all_sample_ids)
    validation_occurrences = []

    for fold_data in folds:
        training = fold_data['training']
        validation = fold_data['validation']
        training_set = set(training)
        validation_set = set(validation)
        fold = fold_data['fold']

        if not training or not validation:
            raise ValueError(f'Repetition {repetition}, fold {fold} contains an empty split.')
        if len(training) != len(training_set) or len(validation) != len(validation_set):
            raise ValueError(f'Repetition {repetition}, fold {fold} contains duplicate sample IDs.')
        if training_set & validation_set:
            raise ValueError(f'Repetition {repetition}, fold {fold} has training/validation overlap.')
        if training_set | validation_set != all_samples or training_set != all_samples - validation_set:
            raise ValueError(f'Repetition {repetition}, fold {fold} does not form a complete complementary split.')

        validation_occurrences.extend(validation)

    if len(validation_occurrences) != len(all_sample_ids) or set(validation_occurrences) != all_samples:
        raise ValueError(f'Repetition {repetition} does not use every sample as validation exactly once.')


def create_repetitions(sample_ids, num_repetitions, num_folds, base_seed):
    """Create deterministic split collections for every repetition."""
    repetitions = []

    for repetition in range(1, int(num_repetitions) + 1):
        seed = int(base_seed) + repetition - 1
        shuffled_ids = list(sample_ids)
        random.Random(seed).shuffle(shuffled_ids)
        folds = create_kfold_splits(shuffled_ids, num_folds)
        validate_kfold_splits(folds, sample_ids, repetition)
        repetitions.append({'repetition': repetition, 'seed': seed, 'folds': folds})

    return repetitions


def prepare_output_dir(output_dir, clean_output_dir):
    """Remove only artefacts managed by this generator."""
    output_dir = Path(output_dir)

    if clean_output_dir and output_dir.exists():
        shutil.rmtree(output_dir)

    output_dir.mkdir(exist_ok=True, parents=True)

    managed_files = {SUMMARY_FILE_NAME, MEMBERSHIP_FILE_NAME, TEST_CASES_WORKBOOK_NAME, *LEGACY_SUMMARY_FILE_NAMES}

    for child in list(output_dir.iterdir()):
        if child.is_dir() and REPETITION_DIR_PATTERN.fullmatch(child.name):
            shutil.rmtree(child)
        elif child.is_file() and (child.name in managed_files or LEGACY_FOLD_FILE_PATTERN.fullmatch(child.name)):
            child.unlink()


def sorted_for_output(sample_ids):
    """Return naturally sorted output when configured."""
    return sorted(sample_ids, key=natural_key) if SORT_OUTPUT_FILES else sample_ids


def write_repetition_files(output_dir, repetitions):
    """Write training_fN.txt and val_fN.txt beneath each repetition directory."""
    for repetition_data in repetitions:
        repetition_dir = Path(output_dir) / f'repetition_{repetition_data["repetition"]}'
        repetition_dir.mkdir(exist_ok=False, parents=True)

        for fold_data in repetition_data['folds']:
            for split_name, prefix in (('training', 'training'), ('validation', 'val')):
                values = '\n'.join(sorted_for_output(fold_data[split_name]))
                (repetition_dir / f'{prefix}_f{fold_data["fold"]}.txt').write_text(f'{values}\n', encoding='utf-8')


def write_test_cases_workbook(output_dir, test_sample_ids):
    """Write a manifest for the fixed external test cohort."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = TEST_CASES_SHEET_NAME
    sheet.append(['sample_id', 'dataset_split', 'included_in_repeated_kfold'])

    for sample_id in sorted(test_sample_ids, key=natural_key):
        sheet.append([sample_id, 'test', False])

    fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:C{sheet.max_row}'
    sheet.column_dimensions['A'].width = max(12, max(len(value) for value in test_sample_ids) + 2)
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 29
    sheet.sheet_view.showGridLines = False
    workbook.save(Path(output_dir) / TEST_CASES_WORKBOOK_NAME)


def write_summary_csv(output_dir, repetitions, source_count, cross_validation_count, test_count):
    """Write repeated k-fold counts and fractions."""
    with open(Path(output_dir) / SUMMARY_FILE_NAME, 'w', newline='', encoding='utf-8') as output:
        writer = csv.writer(output)
        writer.writerow(['repetition', 'repetition_seed', 'fold', 'source_count', 'held_out_test_count', 'cross_validation_count',
                         'training_count', 'validation_count', 'training_fraction', 'validation_fraction'])
        for repetition_data in repetitions:
            for fold_data in repetition_data['folds']:
                training_count = len(fold_data['training'])
                validation_count = len(fold_data['validation'])
                writer.writerow([repetition_data['repetition'], repetition_data['seed'], fold_data['fold'], source_count, test_count,
                                 cross_validation_count, training_count, validation_count,
                                 round(training_count / cross_validation_count, 4), round(validation_count / cross_validation_count, 4)])


def write_membership_csv(output_dir, repetitions):
    """Write every repeated k-fold assignment in long format."""
    with open(Path(output_dir) / MEMBERSHIP_FILE_NAME, 'w', newline='', encoding='utf-8') as output:
        writer = csv.writer(output)
        writer.writerow(['repetition', 'repetition_seed', 'fold', 'split', 'sample_id'])
        for repetition_data in repetitions:
            for fold_data in repetition_data['folds']:
                for split_name in ('training', 'validation'):
                    for sample_id in sorted_for_output(fold_data[split_name]):
                        writer.writerow([repetition_data['repetition'], repetition_data['seed'], fold_data['fold'], split_name, sample_id])


def print_summary(output_dir, repetitions, source_count, cross_validation_count, test_count, base_seed):
    """Print the resolved repeated k-fold collection."""
    print('======================================================================================')
    print(f'Repeated k-fold output directory: {output_dir}')
    print(f'Source samples: {source_count}; held-out test: {test_count}; cross-validation: {cross_validation_count}')
    print(f'Base seed: {base_seed}; repetitions: {len(repetitions)}; folds: {len(repetitions[0]["folds"])}')
    for repetition_data in repetitions:
        print(f'Repetition {repetition_data["repetition"]} (seed={repetition_data["seed"]})')
        for fold_data in repetition_data['folds']:
            print(f'  Fold {fold_data["fold"]}: training={len(fold_data["training"])}; validation={len(fold_data["validation"])}')
    print('======================================================================================')


def create_repeated_kfold_lists(mark_list_path, output_dir, num_repetitions, num_folds, base_seed, test_sample_ids=None):
    """Create repeated k-fold lists after reserving an optional fixed test cohort."""
    all_sample_ids = load_unique_sample_ids(mark_list_path)
    test_sample_ids = normalise_test_sample_ids(test_sample_ids, all_sample_ids, mark_list_path)
    test_set = set(test_sample_ids)
    cross_validation_ids = [sample_id for sample_id in all_sample_ids if sample_id not in test_set]
    validate_generation_args(len(cross_validation_ids), num_repetitions, num_folds)
    repetitions = create_repetitions(cross_validation_ids, num_repetitions, num_folds, base_seed)

    prepare_output_dir(output_dir, CLEAN_OUTPUT_DIR)
    write_repetition_files(output_dir, repetitions)

    if test_sample_ids:
        write_test_cases_workbook(output_dir, test_sample_ids)
    if WRITE_SUMMARY_CSV:
        write_summary_csv(output_dir, repetitions, len(all_sample_ids), len(cross_validation_ids), len(test_sample_ids))
    if WRITE_MEMBERSHIP_CSV:
        write_membership_csv(output_dir, repetitions)

    print_summary(output_dir, repetitions, len(all_sample_ids), len(cross_validation_ids), len(test_sample_ids), base_seed)
    return repetitions


def main():
    """Run repeated k-fold generation using the configuration block."""
    create_repeated_kfold_lists(MARK_LIST_PATH, OUTPUT_DIR, NUM_REPETITIONS, NUM_FOLDS_PER_REPETITION, BASE_SEED, TEST_SAMPLE_IDS)


if __name__ == '__main__':
    main()
