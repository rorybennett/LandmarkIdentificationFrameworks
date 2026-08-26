"""Regression tests for IPV repeated k-fold generation and validation."""

import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from IPV.utils.fold_utils import calculate_fold_collection_sha256, validate_repeated_kfold_lists
from IPV.utils.generate_folds import (MEMBERSHIP_FILE_NAME, SUMMARY_FILE_NAME, TEST_CASES_WORKBOOK_NAME,
                                      create_repeated_kfold_lists)


class RepeatedKFoldTests(unittest.TestCase):
    @staticmethod
    def make_mark_list(directory, sample_count=13):
        path = Path(directory) / 'marks.txt'
        path.write_text(''.join(f'A{index}.png (1, 2)\n' for index in range(1, sample_count + 1)), encoding='utf-8')
        return path

    @staticmethod
    def read_ids(path):
        return {line.strip() for line in Path(path).read_text(encoding='utf-8').splitlines() if line.strip()}

    def test_every_eligible_sample_is_validation_once_per_repetition(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            mark_list = self.make_mark_list(root)
            output = root / 'folds'
            held_out = {'A2', 'A12'}
            expected = {f'A{index}' for index in range(1, 14)} - held_out
            create_repeated_kfold_lists(mark_list, output, num_repetitions=3, num_folds=4, base_seed=17,
                                        test_sample_ids=sorted(held_out))

            self.assertEqual(validate_repeated_kfold_lists(output),
                             {'repetitions': [1, 2, 3], 'folds': [1, 2, 3, 4], 'sample_count': len(expected)})

            for repetition in range(1, 4):
                validation_occurrences = []
                for fold in range(1, 5):
                    repetition_dir = output / f'repetition_{repetition}'
                    training = self.read_ids(repetition_dir / f'training_f{fold}.txt')
                    validation = self.read_ids(repetition_dir / f'val_f{fold}.txt')
                    self.assertFalse(training & validation)
                    self.assertEqual(training | validation, expected)
                    validation_occurrences.extend(validation)
                self.assertEqual(len(validation_occurrences), len(expected))
                self.assertEqual(set(validation_occurrences), expected)

            workbook = load_workbook(output / TEST_CASES_WORKBOOK_NAME, read_only=True)
            self.assertEqual([row[0] for row in list(workbook.active.iter_rows(values_only=True))[1:]], ['A2', 'A12'])
            workbook.close()

    def test_summary_membership_and_digest_are_deterministic(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            mark_list = self.make_mark_list(root, sample_count=12)
            outputs = [root / 'first', root / 'second']

            for output in outputs:
                create_repeated_kfold_lists(mark_list, output, num_repetitions=2, num_folds=3, base_seed=9)

            first_digest = calculate_fold_collection_sha256(outputs[0])
            second_digest = calculate_fold_collection_sha256(outputs[1])
            self.assertEqual(first_digest, second_digest)

            with open(outputs[0] / SUMMARY_FILE_NAME, newline='', encoding='utf-8') as summary_file:
                self.assertEqual(len(list(csv.DictReader(summary_file))), 6)
            with open(outputs[0] / MEMBERSHIP_FILE_NAME, newline='', encoding='utf-8') as membership_file:
                rows = list(csv.DictReader(membership_file))
            self.assertEqual({row['split'] for row in rows}, {'training', 'validation'})

            target = outputs[1] / 'repetition_1' / 'val_f1.txt'
            target.write_text(target.read_text(encoding='utf-8') + 'changed_sample\n', encoding='utf-8')
            self.assertNotEqual(first_digest, calculate_fold_collection_sha256(outputs[1]))

    def test_invalid_configuration_preserves_existing_outputs(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            mark_list = self.make_mark_list(root, sample_count=5)
            output = root / 'folds'
            sentinel = output / 'repetition_1' / 'sentinel.txt'
            sentinel.parent.mkdir(parents=True)
            sentinel.write_text('preserve', encoding='utf-8')

            with self.assertRaisesRegex(ValueError, 'Not enough samples'):
                create_repeated_kfold_lists(mark_list, output, num_repetitions=1, num_folds=4, base_seed=42,
                                            test_sample_ids=['A1', 'A2'])

            self.assertEqual(sentinel.read_text(encoding='utf-8'), 'preserve')


if __name__ == '__main__':
    unittest.main()
