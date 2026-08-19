"""Regression tests for repeated k-fold list generation."""

import csv
import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


MODULE_PATH = Path(__file__).parents[1] / 'Heatmaps' / 'utils' / 'generate_folds.py'
SPEC = importlib.util.spec_from_file_location('heatmap_generate_folds', MODULE_PATH)
generate_folds = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(generate_folds)


class RepeatedKFoldTests(unittest.TestCase):
    def make_mark_list(self, directory, sample_count=11, prefix='patient_'):
        mark_list = Path(directory) / 'marks.txt'
        rows = [f'{prefix}{index}.png (1, 2)\n' for index in range(1, sample_count + 1)]
        mark_list.write_text(''.join(rows), encoding='utf-8')
        return mark_list

    @staticmethod
    def read_ids(path):
        return {line.strip() for line in Path(path).read_text(encoding='utf-8').splitlines() if line.strip()}

    def test_every_sample_is_validation_once_per_repetition(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            temporary_path = Path(temporary_dir)
            mark_list = self.make_mark_list(temporary_path)
            output_dir = temporary_path / 'folds'
            expected_samples = {f'patient_{index}' for index in range(1, 12)}

            generate_folds.create_repeated_kfold_lists(mark_list_path=mark_list, output_dir=output_dir, num_repetitions=3,
                                                       num_folds=4, base_seed=42)

            self.assertEqual({path.name for path in output_dir.glob('repetition_*')}, {'repetition_1', 'repetition_2', 'repetition_3'})

            for repetition in range(1, 4):
                repetition_dir = output_dir / f'repetition_{repetition}'
                expected_files = {f'{prefix}_f{fold}.txt' for fold in range(1, 5) for prefix in ('training', 'val')}
                self.assertEqual({path.name for path in repetition_dir.iterdir()}, expected_files)
                validation_occurrences = []

                for fold in range(1, 5):
                    training_ids = self.read_ids(repetition_dir / f'training_f{fold}.txt')
                    validation_ids = self.read_ids(repetition_dir / f'val_f{fold}.txt')
                    self.assertFalse(training_ids & validation_ids)
                    self.assertEqual(training_ids | validation_ids, expected_samples)
                    self.assertEqual(training_ids, expected_samples - validation_ids)
                    validation_occurrences.extend(validation_ids)

                self.assertEqual(len(validation_occurrences), len(expected_samples))
                self.assertEqual(set(validation_occurrences), expected_samples)

    def test_test_samples_are_excluded_from_every_repetition_and_recorded_in_excel(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            temporary_path = Path(temporary_dir)
            mark_list = self.make_mark_list(temporary_path, sample_count=50, prefix='A')
            output_dir = temporary_path / 'folds'
            test_sample_ids = ['A4', 'A50', 'A8']
            expected_test_samples = {'A4', 'A8', 'A50'}
            expected_cross_validation_samples = {f'A{index}' for index in range(1, 51)} - expected_test_samples

            generate_folds.create_repeated_kfold_lists(mark_list_path=mark_list, output_dir=output_dir, num_repetitions=3,
                                                       num_folds=5, base_seed=42, test_sample_ids=test_sample_ids)

            for repetition in range(1, 4):
                validation_occurrences = []

                for fold in range(1, 6):
                    repetition_dir = output_dir / f'repetition_{repetition}'
                    training_ids = self.read_ids(repetition_dir / f'training_f{fold}.txt')
                    validation_ids = self.read_ids(repetition_dir / f'val_f{fold}.txt')
                    self.assertFalse((training_ids | validation_ids) & expected_test_samples)
                    self.assertEqual(training_ids | validation_ids, expected_cross_validation_samples)
                    validation_occurrences.extend(validation_ids)

                self.assertEqual(len(validation_occurrences), len(expected_cross_validation_samples))
                self.assertEqual(set(validation_occurrences), expected_cross_validation_samples)

            self.assertFalse(any(path.name.startswith('test_') and path.suffix == '.txt' for path in output_dir.rglob('*')))
            with open(output_dir / generate_folds.MEMBERSHIP_FILE_NAME, 'r', encoding='utf-8') as membership_file:
                membership_rows = list(csv.DictReader(membership_file))
            self.assertEqual({row['split'] for row in membership_rows}, {'training', 'validation'})
            self.assertFalse({row['sample_id'] for row in membership_rows} & expected_test_samples)

            with open(output_dir / generate_folds.SUMMARY_FILE_NAME, 'r', encoding='utf-8') as summary_file:
                summary_rows = list(csv.DictReader(summary_file))
            self.assertEqual({int(row['source_count']) for row in summary_rows}, {50})
            self.assertEqual({int(row['held_out_test_count']) for row in summary_rows}, {3})
            self.assertEqual({int(row['cross_validation_count']) for row in summary_rows}, {47})

            workbook_path = output_dir / generate_folds.TEST_CASES_WORKBOOK_NAME
            workbook = load_workbook(workbook_path)
            self.assertEqual(workbook.sheetnames, [generate_folds.TEST_CASES_SHEET_NAME])
            sheet = workbook[generate_folds.TEST_CASES_SHEET_NAME]
            self.assertEqual(list(sheet.iter_rows(values_only=True)), [
                ('sample_id', 'dataset_split', 'included_in_repeated_kfold'),
                ('A4', 'test', False),
                ('A8', 'test', False),
                ('A50', 'test', False),
            ])
            self.assertEqual(sheet.freeze_panes, 'A2')
            self.assertEqual(sheet.auto_filter.ref, 'A1:C4')
            workbook.close()

    def test_none_and_empty_test_sample_lists_use_the_full_dataset_without_a_workbook(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            temporary_path = Path(temporary_dir)
            mark_list = self.make_mark_list(temporary_path, sample_count=8)
            expected_samples = {f'patient_{index}' for index in range(1, 9)}

            for output_name, test_sample_ids in (('none', None), ('empty', [])):
                with self.subTest(test_sample_ids=test_sample_ids):
                    output_dir = temporary_path / output_name
                    generate_folds.create_repeated_kfold_lists(mark_list_path=mark_list, output_dir=output_dir, num_repetitions=1,
                                                               num_folds=4, base_seed=42, test_sample_ids=test_sample_ids)
                    self.assertFalse((output_dir / generate_folds.TEST_CASES_WORKBOOK_NAME).exists())
                    self.assertEqual(self.read_ids(output_dir / 'repetition_1' / 'training_f1.txt') |
                                     self.read_ids(output_dir / 'repetition_1' / 'val_f1.txt'), expected_samples)

    def test_dotted_sample_stem_can_be_excluded_by_exact_id_or_filename(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            temporary_path = Path(temporary_dir)
            mark_list = temporary_path / 'marks.txt'
            mark_list.write_text('case.v1.png (1, 2)\ncase2.png (1, 2)\ncase3.png (1, 2)\n', encoding='utf-8')

            for output_name, test_sample_id in (('stem', 'case.v1'), ('filename', 'case.v1.png')):
                with self.subTest(test_sample_id=test_sample_id):
                    output_dir = temporary_path / output_name
                    generate_folds.create_repeated_kfold_lists(mark_list_path=mark_list, output_dir=output_dir, num_repetitions=1,
                                                               num_folds=2, base_seed=42, test_sample_ids=[test_sample_id])
                    fold_ids = (
                        self.read_ids(output_dir / 'repetition_1' / 'training_f1.txt') |
                        self.read_ids(output_dir / 'repetition_1' / 'val_f1.txt')
                    )
                    self.assertEqual(fold_ids, {'case2', 'case3'})

                    workbook = load_workbook(output_dir / generate_folds.TEST_CASES_WORKBOOK_NAME, read_only=True)
                    self.assertEqual(workbook[generate_folds.TEST_CASES_SHEET_NAME]['A2'].value, 'case.v1')
                    workbook.close()

    def test_generation_is_deterministic(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            temporary_path = Path(temporary_dir)
            mark_list = self.make_mark_list(temporary_path)
            first_output = temporary_path / 'first'
            second_output = temporary_path / 'second'

            for output_dir in (first_output, second_output):
                generate_folds.create_repeated_kfold_lists(mark_list_path=mark_list, output_dir=output_dir, num_repetitions=2,
                                                           num_folds=5, base_seed=17)

            first_files = {path.relative_to(first_output): path.read_text(encoding='utf-8') for path in first_output.rglob('*') if path.is_file()}
            second_files = {path.relative_to(second_output): path.read_text(encoding='utf-8') for path in second_output.rglob('*') if path.is_file()}
            self.assertEqual(first_files, second_files)

    def test_regeneration_removes_legacy_flat_fold_files_only(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            temporary_path = Path(temporary_dir)
            mark_list = self.make_mark_list(temporary_path)
            output_dir = temporary_path / 'folds'
            output_dir.mkdir()
            legacy_names = ('train_f1.txt', 'training_f1.txt', 'val_f1.txt', 'obsolete_f1.txt', 'fold_summary.csv', 'fold_membership.csv')

            for legacy_name in legacy_names:
                (output_dir / legacy_name).write_text('legacy\n', encoding='utf-8')

            unrelated_file = output_dir / 'notes.txt'
            unrelated_file.write_text('keep me\n', encoding='utf-8')
            generate_folds.create_repeated_kfold_lists(mark_list_path=mark_list, output_dir=output_dir, num_repetitions=1,
                                                       num_folds=3, base_seed=42)

            self.assertFalse(any((output_dir / legacy_name).exists() for legacy_name in legacy_names))
            self.assertEqual(unrelated_file.read_text(encoding='utf-8'), 'keep me\n')

    def test_invalid_test_sample_configuration_preserves_existing_outputs(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            temporary_path = Path(temporary_dir)
            mark_list = self.make_mark_list(temporary_path, sample_count=5)
            output_dir = temporary_path / 'folds'
            repetition_dir = output_dir / 'repetition_1'
            repetition_dir.mkdir(parents=True)
            sentinel = repetition_dir / 'sentinel.txt'
            sentinel.write_text('keep me\n', encoding='utf-8')

            invalid_cases = (
                (['missing_patient'], 'not found in mark-list file'),
                (['patient_1', 'patient_1.png'], 'Duplicate sample IDs'),
                ('patient_1', 'not one bare string'),
                ([''], 'entry 1 is blank'),
                (['patient_1', 'patient_2', 'patient_3'], 'Not enough samples'),
            )

            for test_sample_ids, expected_error in invalid_cases:
                with self.subTest(test_sample_ids=test_sample_ids):
                    with self.assertRaisesRegex(ValueError, expected_error):
                        generate_folds.create_repeated_kfold_lists(mark_list_path=mark_list, output_dir=output_dir, num_repetitions=1,
                                                                   num_folds=3, base_seed=42, test_sample_ids=test_sample_ids)
                    self.assertEqual(sentinel.read_text(encoding='utf-8'), 'keep me\n')

    def test_regeneration_without_test_samples_removes_stale_test_manifest_only(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            temporary_path = Path(temporary_dir)
            mark_list = self.make_mark_list(temporary_path, sample_count=8)
            output_dir = temporary_path / 'folds'
            generate_folds.create_repeated_kfold_lists(mark_list_path=mark_list, output_dir=output_dir, num_repetitions=1,
                                                       num_folds=4, base_seed=42, test_sample_ids=['patient_1'])
            workbook_path = output_dir / generate_folds.TEST_CASES_WORKBOOK_NAME
            self.assertTrue(workbook_path.is_file())
            unrelated_file = output_dir / 'notes.txt'
            unrelated_file.write_text('keep me\n', encoding='utf-8')

            generate_folds.create_repeated_kfold_lists(mark_list_path=mark_list, output_dir=output_dir, num_repetitions=1,
                                                       num_folds=4, base_seed=42, test_sample_ids=[])

            self.assertFalse(workbook_path.exists())
            self.assertEqual(unrelated_file.read_text(encoding='utf-8'), 'keep me\n')

    def test_invalid_repetition_and_fold_counts_are_rejected(self):
        with self.assertRaisesRegex(ValueError, 'num_repetitions must be at least 1'):
            generate_folds.validate_generation_args(sample_count=10, num_repetitions=0, num_folds=5)

        with self.assertRaisesRegex(ValueError, 'num_folds must be at least 2'):
            generate_folds.validate_generation_args(sample_count=10, num_repetitions=1, num_folds=1)

        with self.assertRaisesRegex(ValueError, 'Not enough samples'):
            generate_folds.validate_generation_args(sample_count=4, num_repetitions=1, num_folds=5)


if __name__ == '__main__':
    unittest.main()
