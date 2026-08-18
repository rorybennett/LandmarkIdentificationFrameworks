"""
Create deterministic repeated k-fold training and validation lists from one mark-list file.

An optional fixed test cohort can be excluded before cross-validation. Every repetition independently shuffles the
remaining dataset, divides it into balanced validation folds, and uses all other eligible samples for training. The
test cohort is recorded in an Excel manifest but is not written as a training-pipeline split.
"""

import csv
import random
import re
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


NUM_REPETITIONS = 3
NUM_FOLDS_PER_REPETITION = 5
BASE_SEED = 42

# Leave as [] or None to use every mark-list sample in repeated k-fold cross-validation.
# Example: ['A4', 'A50', 'A8']
TEST_SAMPLE_IDS = []

MARK_LIST_PATH = Path(r'C:\Storage\Datasets\IPV\OriginalData\doctors_resampled_transverseMarkList.txt')
OUTPUT_DIR = Path(r'C:\Storage\Datasets\IPV\OriginalData\folds_network_study')

CLEAN_OUTPUT_DIR = False
SORT_OUTPUT_FILES = True
WRITE_SUMMARY_CSV = True
WRITE_MEMBERSHIP_CSV = True

REPETITION_DIR_PREFIX = 'repetition_'
TRAINING_PREFIX = 'training'
VALIDATION_PREFIX = 'val'
SUMMARY_FILE_NAME = 'repeated_kfold_summary.csv'
MEMBERSHIP_FILE_NAME = 'repeated_kfold_membership.csv'
TEST_CASES_WORKBOOK_NAME = 'test_cases.xlsx'
TEST_CASES_SHEET_NAME = 'test_cases'
REPETITION_DIR_PATTERN = re.compile(r'^repetition_(\d+)$')
LEGACY_FOLD_FILE_PATTERN = re.compile(r'^[A-Za-z]+_f\d+\.txt$')
LEGACY_SUMMARY_FILE_NAMES = ('fold_summary.csv', 'fold_membership.csv')


def natural_key(value):
    """Create a natural sorting key, so A2 comes before A10."""
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r'(\d+)', str(value))]


def read_sample_ids(mark_list_path):
    """Read sample IDs from the first column of a mark-list file."""
    mark_list_path = Path(mark_list_path)

    if not mark_list_path.is_file():
        raise FileNotFoundError(f'Mark-list file not found: {mark_list_path}')

    sample_ids = []

    with open(mark_list_path, 'r', encoding='utf-8') as mark_file:
        for line in mark_file:
            line = line.strip()

            if not line:
                continue

            image_name = line.split()[0]
            sample_ids.append(Path(image_name).stem)

    return sample_ids


def check_duplicates(sample_ids, source_name):
    """Stop execution if duplicated sample IDs are present."""
    seen = set()
    duplicates = set()

    for sample_id in sample_ids:
        if sample_id in seen:
            duplicates.add(sample_id)

        seen.add(sample_id)

    if duplicates:
        duplicate_text = ', '.join(sorted(duplicates, key=natural_key))
        raise ValueError(f'Duplicate sample IDs found in {source_name}: {duplicate_text}')


def load_unique_sample_ids(mark_list_path):
    """Load, validate, and naturally sort sample IDs before shuffling."""
    sample_ids = read_sample_ids(mark_list_path)
    check_duplicates(sample_ids, str(mark_list_path))

    if not sample_ids:
        raise ValueError(f'No sample IDs found in {mark_list_path}')

    return sorted(sample_ids, key=natural_key)


def normalise_test_sample_ids(test_sample_ids, all_sample_ids, mark_list_path):
    """Validate optional held-out test IDs and return naturally sorted mark-list stems."""
    if test_sample_ids is None:
        return []

    if isinstance(test_sample_ids, (str, Path)):
        raise ValueError('TEST_SAMPLE_IDS must be an iterable of sample IDs, not one bare string or path.')

    try:
        raw_sample_ids = list(test_sample_ids)
    except TypeError as error:
        raise ValueError('TEST_SAMPLE_IDS must be None or an iterable of sample ID strings.') from error

    available_sample_set = set(all_sample_ids)
    normalised_sample_ids = []

    for index, raw_sample_id in enumerate(raw_sample_ids, start=1):
        if not isinstance(raw_sample_id, (str, Path)):
            raise ValueError(f'TEST_SAMPLE_IDS entry {index} must be a string or path. Got: {raw_sample_id!r}')

        sample_text = str(raw_sample_id).strip()

        if not sample_text:
            raise ValueError(f'TEST_SAMPLE_IDS entry {index} is blank.')

        sample_id = sample_text if sample_text in available_sample_set else Path(sample_text).stem

        if not sample_id:
            raise ValueError(f'TEST_SAMPLE_IDS entry {index} does not contain a valid sample ID: {raw_sample_id!r}')

        normalised_sample_ids.append(sample_id)

    check_duplicates(normalised_sample_ids, 'TEST_SAMPLE_IDS')
    unknown_sample_ids = sorted(set(normalised_sample_ids) - available_sample_set, key=natural_key)

    if unknown_sample_ids:
        unknown_text = ', '.join(unknown_sample_ids)
        raise ValueError(f'Test sample IDs were not found in mark-list file {mark_list_path}: {unknown_text}')

    return sorted(normalised_sample_ids, key=natural_key)


def validate_generation_args(sample_count, num_repetitions, num_folds):
    """Validate repeated k-fold configuration values."""
    if int(num_repetitions) < 1:
        raise ValueError(f'num_repetitions must be at least 1. Got: {num_repetitions}')

    if int(num_folds) < 2:
        raise ValueError(f'num_folds must be at least 2 for k-fold cross-validation. Got: {num_folds}')

    if int(sample_count) < int(num_folds):
        raise ValueError(f'Not enough samples for {num_folds} folds. Found {sample_count} samples; at least {num_folds} are required.')


def make_balanced_chunks(sample_ids, num_chunks):
    """Split shuffled sample IDs into balanced chunks."""
    chunks = [[] for _ in range(int(num_chunks))]

    for index, sample_id in enumerate(sample_ids):
        chunks[index % int(num_chunks)].append(sample_id)

    return chunks


def create_kfold_splits(sample_ids, num_folds):
    """Create one complete k-fold training/validation split collection."""
    validation_chunks = make_balanced_chunks(sample_ids, num_chunks=num_folds)
    folds = []

    for fold_index, validation_ids in enumerate(validation_chunks, start=1):
        validation_set = set(validation_ids)
        training_ids = [sample_id for sample_id in sample_ids if sample_id not in validation_set]
        folds.append({'fold': fold_index, 'training': training_ids, 'validation': validation_ids})

    return folds


def validate_kfold_splits(folds, all_sample_ids, repetition):
    """Validate one repetition and ensure every sample is validation exactly once."""
    all_sample_set = set(all_sample_ids)
    validation_occurrences = []

    for fold in folds:
        training_ids = fold['training']
        validation_ids = fold['validation']
        training_set = set(training_ids)
        validation_set = set(validation_ids)
        fold_number = fold['fold']

        if not training_ids:
            raise ValueError(f'Repetition {repetition}, fold {fold_number} has an empty training split.')

        if not validation_ids:
            raise ValueError(f'Repetition {repetition}, fold {fold_number} has an empty validation split.')

        if len(training_ids) != len(training_set):
            raise ValueError(f'Repetition {repetition}, fold {fold_number} training split contains duplicate sample IDs.')

        if len(validation_ids) != len(validation_set):
            raise ValueError(f'Repetition {repetition}, fold {fold_number} validation split contains duplicate sample IDs.')

        overlap = training_set & validation_set

        if overlap:
            overlap_text = ', '.join(sorted(overlap, key=natural_key))
            raise ValueError(f'Repetition {repetition}, fold {fold_number} has training/validation overlap: {overlap_text}')

        if training_set | validation_set != all_sample_set:
            raise ValueError(f'Repetition {repetition}, fold {fold_number} does not cover the full cross-validation-eligible dataset.')

        if training_set != all_sample_set - validation_set:
            raise ValueError(f'Repetition {repetition}, fold {fold_number} training split is not the complement of its validation split.')

        validation_occurrences.extend(validation_ids)

    if len(validation_occurrences) != len(all_sample_ids) or set(validation_occurrences) != all_sample_set:
        raise ValueError(f'Repetition {repetition} does not use every sample as validation exactly once.')

    if len(validation_occurrences) != len(set(validation_occurrences)):
        raise ValueError(f'Repetition {repetition} uses at least one sample as validation in more than one fold.')


def create_repetitions(sample_ids, num_repetitions, num_folds, base_seed):
    """Create deterministic k-fold collections for every repetition."""
    repetitions = []

    for repetition in range(1, int(num_repetitions) + 1):
        repetition_seed = int(base_seed) + repetition - 1
        shuffled_ids = list(sample_ids)
        random.Random(repetition_seed).shuffle(shuffled_ids)
        folds = create_kfold_splits(shuffled_ids, num_folds=num_folds)
        validate_kfold_splits(folds=folds, all_sample_ids=sample_ids, repetition=repetition)
        repetitions.append({'repetition': repetition, 'seed': repetition_seed, 'folds': folds})

    return repetitions


def validate_test_sample_exclusion(repetitions, test_sample_ids):
    """Ensure held-out test IDs do not occur in any generated training or validation split."""
    test_sample_set = set(test_sample_ids)

    if not test_sample_set:
        return

    for repetition_data in repetitions:
        for fold in repetition_data['folds']:
            split_sample_set = set(fold['training']) | set(fold['validation'])
            leaked_sample_ids = split_sample_set & test_sample_set

            if leaked_sample_ids:
                leaked_text = ', '.join(sorted(leaked_sample_ids, key=natural_key))
                raise ValueError(f'Held-out test sample IDs leaked into repetition {repetition_data["repetition"]}, '
                                 f'fold {fold["fold"]}: {leaked_text}')


def prepare_output_dir(output_dir, clean_output_dir):
    """Prepare the root and remove stale fold artefacts managed by this generator."""
    output_dir = Path(output_dir)

    if clean_output_dir and output_dir.exists():
        shutil.rmtree(output_dir)

    output_dir.mkdir(exist_ok=True, parents=True)

    for child in output_dir.iterdir():
        if child.is_dir() and REPETITION_DIR_PATTERN.fullmatch(child.name):
            shutil.rmtree(child)

    for managed_file_name in (SUMMARY_FILE_NAME, MEMBERSHIP_FILE_NAME, TEST_CASES_WORKBOOK_NAME):
        managed_file_path = output_dir / managed_file_name

        if managed_file_path.exists():
            managed_file_path.unlink()

    for child in output_dir.iterdir():
        if child.is_file() and (LEGACY_FOLD_FILE_PATTERN.fullmatch(child.name) or child.name in LEGACY_SUMMARY_FILE_NAMES):
            child.unlink()


def sorted_for_output(sample_ids):
    """Sort output files if configured to do so."""
    return sorted(sample_ids, key=natural_key) if SORT_OUTPUT_FILES else sample_ids


def write_sample_list(path, sample_ids):
    """Write one sample ID per line."""
    with open(path, 'w', encoding='utf-8') as output_file:
        for sample_id in sorted_for_output(sample_ids):
            output_file.write(f'{sample_id}\n')


def write_repetition_files(output_dir, repetitions):
    """Write training_fN.txt and val_fN.txt files beneath every repetition directory."""
    for repetition_data in repetitions:
        repetition = repetition_data['repetition']
        repetition_dir = Path(output_dir) / f'{REPETITION_DIR_PREFIX}{repetition}'
        repetition_dir.mkdir(exist_ok=False, parents=True)

        for fold in repetition_data['folds']:
            fold_number = fold['fold']
            write_sample_list(repetition_dir / f'{TRAINING_PREFIX}_f{fold_number}.txt', fold['training'])
            write_sample_list(repetition_dir / f'{VALIDATION_PREFIX}_f{fold_number}.txt', fold['validation'])


def write_test_cases_workbook(output_dir, test_sample_ids):
    """Write a root-level Excel manifest for samples excluded from repeated k-fold cross-validation."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = TEST_CASES_SHEET_NAME
    sheet.append(['sample_id', 'dataset_split', 'included_in_repeated_kfold'])

    for sample_id in sorted(test_sample_ids, key=natural_key):
        sheet.append([sample_id, 'test', False])

    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')

    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:C{sheet.max_row}'
    sheet.column_dimensions['A'].width = max(12, max(len(sample_id) for sample_id in test_sample_ids) + 2)
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 29
    sheet.sheet_view.showGridLines = False
    workbook.save(Path(output_dir) / TEST_CASES_WORKBOOK_NAME)


def write_summary_csv(output_dir, repetitions, source_count, cross_validation_count, test_count):
    """Write repeated k-fold counts and fractions."""
    summary_path = Path(output_dir) / SUMMARY_FILE_NAME

    with open(summary_path, 'w', newline='', encoding='utf-8') as summary_file:
        writer = csv.writer(summary_file)
        writer.writerow(['repetition', 'repetition_seed', 'fold', 'source_count', 'held_out_test_count', 'cross_validation_count',
                         'training_count', 'validation_count', 'training_fraction', 'validation_fraction'])

        for repetition_data in repetitions:
            for fold in repetition_data['folds']:
                training_count = len(fold['training'])
                validation_count = len(fold['validation'])
                writer.writerow([
                    repetition_data['repetition'], repetition_data['seed'], fold['fold'], source_count, test_count, cross_validation_count,
                    training_count, validation_count, round(training_count / cross_validation_count, 4),
                    round(validation_count / cross_validation_count, 4),
                ])


def write_membership_csv(output_dir, repetitions):
    """Write every repeated k-fold sample assignment in long format."""
    membership_path = Path(output_dir) / MEMBERSHIP_FILE_NAME

    with open(membership_path, 'w', newline='', encoding='utf-8') as membership_file:
        writer = csv.writer(membership_file)
        writer.writerow(['repetition', 'repetition_seed', 'fold', 'split', 'sample_id'])

        for repetition_data in repetitions:
            for fold in repetition_data['folds']:
                for split_name in ('training', 'validation'):
                    for sample_id in sorted_for_output(fold[split_name]):
                        writer.writerow([repetition_data['repetition'], repetition_data['seed'], fold['fold'], split_name, sample_id])


def print_summary(output_dir, repetitions, source_count, cross_validation_count, test_count, base_seed):
    """Print repeated k-fold split counts to the terminal."""
    print('======================================================================================')
    print(f'Repeated k-fold output directory: {output_dir}')
    print(f'Source samples: {source_count}')
    print(f'Held-out test samples: {test_count}')
    print(f'Cross-validation samples: {cross_validation_count}')
    print(f'Base seed: {base_seed}')
    print(f'Repetitions: {len(repetitions)}')
    print(f'Folds per repetition: {len(repetitions[0]["folds"])}')
    print('--------------------------------------------------------------------------------------')

    for repetition_data in repetitions:
        print(f'Repetition {repetition_data["repetition"]} (seed={repetition_data["seed"]})')

        for fold in repetition_data['folds']:
            training_count = len(fold['training'])
            validation_count = len(fold['validation'])
            print(
                f'  Fold {fold["fold"]}: '
                f'training={training_count} ({training_count / cross_validation_count:.2%}), '
                f'validation={validation_count} ({validation_count / cross_validation_count:.2%})'
            )

    print('======================================================================================')


def create_repeated_kfold_lists(mark_list_path, output_dir, num_repetitions, num_folds, base_seed, test_sample_ids=None):
    """Create repeated k-fold lists after optionally reserving a fixed external test cohort."""
    all_sample_ids = load_unique_sample_ids(mark_list_path)
    test_sample_ids = normalise_test_sample_ids(test_sample_ids=test_sample_ids, all_sample_ids=all_sample_ids,
                                                mark_list_path=mark_list_path)
    test_sample_set = set(test_sample_ids)
    cross_validation_sample_ids = [sample_id for sample_id in all_sample_ids if sample_id not in test_sample_set]
    validate_generation_args(sample_count=len(cross_validation_sample_ids), num_repetitions=num_repetitions, num_folds=num_folds)
    repetitions = create_repetitions(sample_ids=cross_validation_sample_ids, num_repetitions=num_repetitions, num_folds=num_folds,
                                     base_seed=base_seed)
    validate_test_sample_exclusion(repetitions=repetitions, test_sample_ids=test_sample_ids)

    prepare_output_dir(output_dir=output_dir, clean_output_dir=CLEAN_OUTPUT_DIR)
    write_repetition_files(output_dir=output_dir, repetitions=repetitions)

    if test_sample_ids:
        write_test_cases_workbook(output_dir=output_dir, test_sample_ids=test_sample_ids)

    if WRITE_SUMMARY_CSV:
        write_summary_csv(output_dir=output_dir, repetitions=repetitions, source_count=len(all_sample_ids),
                          cross_validation_count=len(cross_validation_sample_ids), test_count=len(test_sample_ids))

    if WRITE_MEMBERSHIP_CSV:
        write_membership_csv(output_dir=output_dir, repetitions=repetitions)

    print_summary(output_dir=output_dir, repetitions=repetitions, source_count=len(all_sample_ids),
                  cross_validation_count=len(cross_validation_sample_ids), test_count=len(test_sample_ids), base_seed=base_seed)
    return repetitions


def main():
    """Run repeated k-fold generation using the configuration block."""
    create_repeated_kfold_lists(mark_list_path=MARK_LIST_PATH, output_dir=OUTPUT_DIR, num_repetitions=NUM_REPETITIONS,
                                num_folds=NUM_FOLDS_PER_REPETITION, base_seed=BASE_SEED, test_sample_ids=TEST_SAMPLE_IDS)


if __name__ == '__main__':
    main()
